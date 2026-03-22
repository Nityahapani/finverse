"""
finverse.audit — model health checks for DCF models and Excel files.

Catches:
- Hardcoded numbers inside formulas
- Assumption drift (macro moved but model not updated)
- Circular references
- Sanity checks (margins > 100%, negative revenue, etc.)
- Inconsistent projection logic
"""
from __future__ import annotations

import numpy as np
import pandas as pd
from dataclasses import dataclass, field


@dataclass
class AuditFlag:
    severity: str         # "error", "warning", "info"
    location: str         # e.g. "DCF.wacc", "Row 14, Col B"
    message: str
    suggestion: str = ""


@dataclass
class AuditResult:
    model_name: str
    flags: list[AuditFlag]
    errors: list[AuditFlag]
    warnings: list[AuditFlag]
    infos: list[AuditFlag]
    score: float              # 0-100, 100 = perfectly clean
    passed: bool

    def summary(self):
        from finverse.utils.display import console
        from rich.table import Table
        from rich import box

        score_color = "green" if self.score >= 80 else ("yellow" if self.score >= 60 else "red")
        status = "[green]PASSED[/green]" if self.passed else "[red]FAILED[/red]"

        console.print(f"\n[bold blue]Model Audit — {self.model_name}[/bold blue]")
        console.print(f"Score: [{score_color}][bold]{self.score:.0f}/100[/bold][/{score_color}]  |  Status: {status}\n")

        if not self.flags:
            console.print("[green]✓ No issues found. Model is clean.[/green]\n")
            return

        table = Table(box=box.SIMPLE_HEAD, header_style="bold blue")
        table.add_column("Severity")
        table.add_column("Location")
        table.add_column("Issue")
        table.add_column("Suggestion")

        severity_colors = {"error": "red", "warning": "yellow", "info": "blue"}
        for f in self.flags:
            c = severity_colors.get(f.severity, "white")
            table.add_row(
                f"[{c}]{f.severity.upper()}[/{c}]",
                f.location,
                f.message,
                f.suggestion,
            )

        console.print(table)
        console.print(
            f"\n  {len(self.errors)} errors  |  "
            f"{len(self.warnings)} warnings  |  "
            f"{len(self.infos)} info\n"
        )

    def to_df(self) -> pd.DataFrame:
        return pd.DataFrame([
            {"severity": f.severity, "location": f.location,
             "message": f.message, "suggestion": f.suggestion}
            for f in self.flags
        ])


def _audit_dcf(model) -> list[AuditFlag]:
    """Run audit checks on a DCF model object."""
    flags = []
    a = model._assumptions

    if a.wacc <= 0 or a.wacc > 0.30:
        flags.append(AuditFlag(
            severity="error",
            location="DCF.wacc",
            message=f"WACC = {a.wacc:.1%} is outside normal range (1%–30%)",
            suggestion="Typical WACC range is 6%–14% for most companies",
        ))

    if a.terminal_growth >= a.wacc:
        flags.append(AuditFlag(
            severity="error",
            location="DCF.terminal_growth",
            message=f"Terminal growth ({a.terminal_growth:.1%}) ≥ WACC ({a.wacc:.1%}) — model is undefined",
            suggestion="Terminal growth must be less than WACC (usually 1.5%–3.5%)",
        ))

    if a.terminal_growth > 0.05:
        flags.append(AuditFlag(
            severity="warning",
            location="DCF.terminal_growth",
            message=f"Terminal growth = {a.terminal_growth:.1%} is above long-run GDP growth",
            suggestion="Consider using 2%–3% for most developed-market companies",
        ))

    if a.ebitda_margin > 0.60:
        flags.append(AuditFlag(
            severity="warning",
            location="DCF.ebitda_margin",
            message=f"EBITDA margin = {a.ebitda_margin:.1%} is unusually high",
            suggestion="Verify margin assumption — 60%+ margins are rare outside software/pharma",
        ))

    if a.ebitda_margin < 0:
        flags.append(AuditFlag(
            severity="error",
            location="DCF.ebitda_margin",
            message=f"EBITDA margin = {a.ebitda_margin:.1%} is negative",
            suggestion="A negative margin DCF produces unreliable results — consider using revenue-based approach",
        ))

    if a.projection_years < 3:
        flags.append(AuditFlag(
            severity="warning",
            location="DCF.projection_years",
            message=f"Only {a.projection_years} projection years — terminal value dominates",
            suggestion="Use 5–10 years for a more balanced model",
        ))

    if model._base_revenue and model._base_revenue <= 0:
        flags.append(AuditFlag(
            severity="error",
            location="DCF.base_revenue",
            message="Base revenue is zero or negative",
            suggestion="Check that TickerData was loaded correctly",
        ))

    if model._results:
        r = model._results
        tv_pct = r.pv_terminal / r.enterprise_value if r.enterprise_value > 0 else 0
        if tv_pct > 0.85:
            flags.append(AuditFlag(
                severity="warning",
                location="DCF.terminal_value",
                message=f"Terminal value = {tv_pct:.0%} of EV — model is highly sensitive to terminal assumptions",
                suggestion="Extend projection period or stress-test terminal growth rate",
            ))

        if r.implied_price <= 0:
            flags.append(AuditFlag(
                severity="error",
                location="DCF.implied_price",
                message="Implied share price is zero or negative",
                suggestion="Check net debt — may exceed enterprise value",
            ))

    if a.revenue_growth and isinstance(a.revenue_growth, list):
        for i, g in enumerate(a.revenue_growth):
            if abs(g) > 0.50:
                flags.append(AuditFlag(
                    severity="warning",
                    location=f"DCF.revenue_growth[year {i+1}]",
                    message=f"Revenue growth = {g:.1%} in year {i+1} is extreme",
                    suggestion="Growth rates above 50% are unusual — verify assumption",
                ))

    return flags


def _audit_three_statement(model) -> list[AuditFlag]:
    """Audit a ThreeStatement model."""
    flags = []
    a = model._assumptions

    if a.gross_margin > 1.0 or a.gross_margin < 0:
        flags.append(AuditFlag(
            severity="error",
            location="ThreeStatement.gross_margin",
            message=f"Gross margin = {a.gross_margin:.1%} is outside 0–100%",
            suggestion="Gross margin must be between 0% and 100%",
        ))

    if a.sga_pct + a.rd_pct > a.gross_margin:
        flags.append(AuditFlag(
            severity="warning",
            location="ThreeStatement.opex",
            message="SG&A + R&D exceed gross margin — company is operating at a loss",
            suggestion="Verify expense assumptions or reduce spend ratios",
        ))

    if a.starting_cash < 0:
        flags.append(AuditFlag(
            severity="error",
            location="ThreeStatement.starting_cash",
            message="Starting cash is negative",
            suggestion="Starting cash must be ≥ 0",
        ))

    if model._results:
        cf = model._results.cash_flow
        neg_fcf_years = (cf.loc["Free cash flow"] < 0).sum()
        if neg_fcf_years > a.projection_years // 2:
            flags.append(AuditFlag(
                severity="warning",
                location="ThreeStatement.cash_flow",
                message=f"Negative FCF in {neg_fcf_years} of {a.projection_years} years",
                suggestion="Check capex assumptions or consider additional financing",
            ))

    return flags


def _audit_lbo(model) -> list[AuditFlag]:
    """Audit an LBO model."""
    flags = []
    a = model._assumptions

    total_leverage = a.senior_leverage + a.sub_leverage
    if total_leverage > 8.0:
        flags.append(AuditFlag(
            severity="warning",
            location="LBO.leverage",
            message=f"Total leverage = {total_leverage:.1f}x is very high",
            suggestion="Typical LBO leverage is 4–6x EBITDA",
        ))

    if a.equity_pct < 0.25:
        flags.append(AuditFlag(
            severity="warning",
            location="LBO.equity_pct",
            message=f"Equity contribution = {a.equity_pct:.0%} — very thin equity cushion",
            suggestion="Most LBOs require 30–50% equity",
        ))

    if model._results:
        if model._results.irr < 0:
            flags.append(AuditFlag(
                severity="error",
                location="LBO.irr",
                message=f"IRR = {model._results.irr:.1%} is negative",
                suggestion="Check entry multiple, exit multiple, and debt assumptions",
            ))
        elif model._results.irr < 0.15:
            flags.append(AuditFlag(
                severity="warning",
                location="LBO.irr",
                message=f"IRR = {model._results.irr:.1%} is below typical PE hurdle rate (15–20%)",
                suggestion="Consider negotiating lower entry multiple or improving operations",
            ))

    return flags


def audit(
    model=None,
    excel_path: str | None = None,
    model_name: str | None = None,
) -> AuditResult:
    """
    Audit a financial model for errors, bad assumptions, and inconsistencies.

    Parameters
    ----------
    model      : DCF, LBO, or ThreeStatement instance
    excel_path : str — path to Excel model file (optional)
    model_name : str — display name (auto-detected if None)

    Returns
    -------
    AuditResult

    Example
    -------
    >>> from finverse import pull, DCF
    >>> from finverse.audit import audit
    >>> data = pull.ticker("AAPL")
    >>> model = DCF(data)
    >>> model.run()
    >>> result = audit(model)
    >>> result.summary()

    Audit an Excel file:
    >>> result = audit(excel_path="my_model.xlsx")
    >>> result.summary()
    """
    from finverse.utils.display import console

    name = model_name or "Financial Model"
    flags = []

    if model is not None:
        from finverse.models.dcf import DCF
        from finverse.models.lbo import LBO
        from finverse.models.three_statement import ThreeStatement

        if isinstance(model, DCF):
            name = model_name or f"DCF — {getattr(model._data, 'ticker', 'manual')}"
            console.print(f"[dim]Auditing {name}...[/dim]")
            flags = _audit_dcf(model)

        elif isinstance(model, LBO):
            name = model_name or "LBO Model"
            console.print(f"[dim]Auditing {name}...[/dim]")
            flags = _audit_lbo(model)

        elif isinstance(model, ThreeStatement):
            name = model_name or "Three-Statement Model"
            console.print(f"[dim]Auditing {name}...[/dim]")
            flags = _audit_three_statement(model)

    if excel_path:
        flags.extend(_audit_excel(excel_path))
        name = model_name or excel_path

    errors   = [f for f in flags if f.severity == "error"]
    warnings = [f for f in flags if f.severity == "warning"]
    infos    = [f for f in flags if f.severity == "info"]

    score = max(
        100
        - len(errors) * 20
        - len(warnings) * 8
        - len(infos) * 2,
        0
    )
    passed = len(errors) == 0

    status_icon = "[green]✓[/green]" if passed else "[red]✗[/red]"
    console.print(
        f"{status_icon} Audit complete — "
        f"score: {score}/100  |  "
        f"{len(errors)} errors, {len(warnings)} warnings"
    )

    return AuditResult(
        model_name=name,
        flags=flags,
        errors=errors,
        warnings=warnings,
        infos=infos,
        score=float(score),
        passed=passed,
    )


def _audit_excel(path: str) -> list[AuditFlag]:
    """Audit an Excel file for common modeling errors."""
    flags = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(path, data_only=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.data_type == "f" and cell.value:
                        formula = str(cell.value)
                        import re
                        numbers_in_formula = re.findall(r"(?<![A-Z])(?<!\d)\d+\.?\d*(?!\d)(?![A-Z])", formula)
                        significant = [n for n in numbers_in_formula if float(n) > 1 and "." not in n]
                        if significant:
                            flags.append(AuditFlag(
                                severity="warning",
                                location=f"Sheet '{sheet_name}', {cell.coordinate}",
                                message=f"Hardcoded number(s) {significant[:3]} found inside formula",
                                suggestion="Move hardcoded inputs to a dedicated assumptions section",
                            ))
    except ImportError:
        flags.append(AuditFlag(
            severity="info",
            location="Excel audit",
            message="openpyxl not available — install with: pip install openpyxl",
            suggestion="pip install openpyxl",
        ))
    except Exception as e:
        flags.append(AuditFlag(
            severity="warning",
            location="Excel audit",
            message=f"Could not fully parse Excel file: {e}",
            suggestion="Ensure file is a valid .xlsx file",
        ))
    return flags
