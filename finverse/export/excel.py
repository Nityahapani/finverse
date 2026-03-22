"""
finverse.export.excel — export DCF results to formatted Excel.
Follows standard banking conventions: blue formulas, green outputs, gray headers.
"""
from __future__ import annotations

from pathlib import Path
import pandas as pd


BLUE   = "FF185FA5"   # formula cells
GREEN  = "FF27500A"   # output cells
GREEN_BG = "FFEAF3DE"
GRAY   = "FF888780"   # section headers
HEADER_BG = "FFE6F1FB"
HEADER_FG = "FF0C447C"
WHITE  = "FFFFFFFF"
BLACK  = "FF1A1A18"


def _get_wb():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import (
            PatternFill, Font, Alignment, Border, Side, numbers
        )
        return Workbook, PatternFill, Font, Alignment, Border, Side
    except ImportError:
        raise ImportError("openpyxl required: pip install openpyxl")


def to_excel(model_or_results, path: str = "finverse_dcf.xlsx") -> str:
    """
    Export a DCF model or results to a formatted Excel file.

    Parameters
    ----------
    model_or_results : DCF instance or DCFResults
    path             : output file path (default "finverse_dcf.xlsx")

    Returns
    -------
    str — absolute path to saved file

    Example
    -------
    >>> from finverse import pull, DCF
    >>> from finverse.export import to_excel
    >>> data = pull.ticker("AAPL")
    >>> model = DCF(data)
    >>> model.run()
    >>> to_excel(model, "apple_dcf.xlsx")
    """
    from finverse.utils.display import console

    Workbook, PatternFill, Font, Alignment, Border, Side = _get_wb()

    from finverse.models.dcf import DCF, DCFResults
    if isinstance(model_or_results, DCF):
        if model_or_results._results is None:
            model_or_results.run()
        results = model_or_results._results
        data = model_or_results._data
    else:
        results = model_or_results
        data = None

    wb = Workbook()

    _write_dcf_sheet(wb.active, results, data, PatternFill, Font, Alignment, Border, Side)

    sens_ws = wb.create_sheet("Sensitivity")
    _write_sensitivity_placeholder(sens_ws, Font, PatternFill, Alignment)

    path = str(Path(path).with_suffix(".xlsx"))
    wb.save(path)
    abs_path = str(Path(path).resolve())
    console.print(f"[green]✓[/green] Saved to [bold]{abs_path}[/bold]")
    return abs_path


def _write_dcf_sheet(ws, results, data, PatternFill, Font, Alignment, Border, Side):
    from openpyxl.utils import get_column_letter

    ws.title = "DCF Model"

    def fill(hex_color):
        return PatternFill("solid", fgColor=hex_color)

    def font(bold=False, color=BLACK, size=11):
        return Font(bold=bold, color=color, size=size, name="Calibri")

    def align(h="left", v="center"):
        return Alignment(horizontal=h, vertical=v)

    def thin_border():
        s = Side(style="thin", color="FFD0CEC8")
        return Border(bottom=s)

    col_widths = [28, 14, 14, 14, 14, 14, 14]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28

    ticker_name = "Financial Model"
    if data and hasattr(data, "name"):
        ticker_name = f"{data.name} — DCF Model"

    ws["A1"] = ticker_name
    ws["A1"].font = Font(bold=True, size=14, color=BLACK, name="Calibri")
    ws["A1"].alignment = align("left", "center")
    ws.row_dimensions[1].height = 30

    a = results.assumptions
    fcf_df = results.fcf_projections
    years = list(fcf_df["year"].astype(int))

    def section_header(row, label):
        ws.cell(row=row, column=1, value=label).font = Font(color=GRAY, size=10, name="Calibri")
        ws.row_dimensions[row].height = 18

    def write_row(row, label, values, is_formula=False, is_output=False, fmt="#,##0.0"):
        ws.cell(row=row, column=1, value=label).font = font(bold=is_output)
        for ci, v in enumerate(values, 2):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.number_format = fmt
            if is_output:
                cell.fill = fill(GREEN_BG)
                cell.font = font(bold=True, color=GREEN)
            elif is_formula:
                cell.font = font(color=BLUE)
            cell.alignment = align("right")

    def write_assumption(row, label, value, fmt="0.0%"):
        ws.cell(row=row, column=1, value=label).font = font()
        cell = ws.cell(row=row, column=2, value=value)
        cell.number_format = fmt
        cell.font = font(bold=True)
        cell.alignment = align("right")

    row = 3
    section_header(row, "── Assumptions ─────────────────────────────")
    row += 1
    write_assumption(row, "WACC", a.wacc); row += 1
    write_assumption(row, "Terminal growth rate", a.terminal_growth); row += 1
    write_assumption(row, "EBITDA margin", a.ebitda_margin); row += 1
    write_assumption(row, "Capex % revenue", a.capex_pct_revenue); row += 1
    write_assumption(row, "Tax rate", a.tax_rate); row += 1

    row += 1
    section_header(row, "── Projections ──────────────────────────────")
    row += 1

    header_row = row
    ws.cell(row=header_row, column=1, value="")
    for ci, yr in enumerate(years, 2):
        cell = ws.cell(row=header_row, column=ci, value=str(yr))
        cell.fill = fill(HEADER_BG)
        cell.font = Font(bold=True, color=HEADER_FG, name="Calibri")
        cell.alignment = align("center")
    row += 1

    write_row(row, "Revenue ($B)", list(fcf_df["revenue"]), is_formula=True); row += 1
    write_row(row, "EBITDA ($B)", list(fcf_df["ebitda"]), is_formula=True); row += 1
    write_row(row, "Free cash flow ($B)", list(fcf_df["fcf"]), is_formula=True); row += 1
    write_row(row, "PV of FCF ($B)", list(fcf_df["pv_fcf"]), is_formula=True); row += 1

    row += 1
    section_header(row, "── Valuation ────────────────────────────────")
    row += 1

    def write_val_row(r, label, value, output=False, fmt="#,##0.0"):
        ws.cell(row=r, column=1, value=label).font = font(bold=output)
        cell = ws.cell(row=r, column=2, value=value)
        cell.number_format = fmt
        if output:
            cell.fill = fill(GREEN_BG)
            cell.font = font(bold=True, color=GREEN)
        cell.alignment = align("right")

    write_val_row(row, "PV of FCFs ($B)", results.pv_fcfs); row += 1
    write_val_row(row, "Terminal value — PV ($B)", results.pv_terminal); row += 1
    write_val_row(row, "Enterprise value ($B)", results.enterprise_value); row += 1
    write_val_row(row, "(-) Net debt ($B)", results.net_debt); row += 1
    write_val_row(row, "Equity value ($B)", results.equity_value, output=True); row += 1
    write_val_row(row, "Implied share price", results.implied_price, output=True, fmt='"$"#,##0.00'); row += 1

    if results.current_price:
        write_val_row(row, "Current price", results.current_price, fmt='"$"#,##0.00'); row += 1
        upside = results.upside_pct or 0
        cell = ws.cell(row=row, column=2, value=upside)
        cell.number_format = "0.0%"
        cell.font = Font(bold=True, color="FF27500A" if upside >= 0 else "FFA32D2D", name="Calibri")
        ws.cell(row=row, column=1, value="Upside / downside").font = font()
        row += 1


def _write_sensitivity_placeholder(ws, Font, PatternFill, Alignment):
    ws["A1"] = "Sensitivity Analysis"
    ws["A1"].font = Font(bold=True, size=13, name="Calibri")
    ws["A2"] = "Run sensitivity() and call .to_excel() to populate this sheet."
    ws["A2"].font = Font(color="FF888780", name="Calibri")
